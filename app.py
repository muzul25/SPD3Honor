import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook

st.set_page_config(page_title="Generator Template SPD Honorium", page_icon="📄")
st.title("📄 Generator Template SPD Honorium")

# Kolom wajib pada database
REQUIRED_COLS = [
    "Nama",
    "Honorarium Persiapan UKOMNAS",
    "Honorarium Pemantauan Briefing UKOMNAS",
    "Honorarium Pelaksanaan UKOMNAS",
    "PPH21",
]

def coerce_numeric(df: pd.DataFrame, cols):
    """Pastikan kolom angka jadi numerik; kosong/teks akan jadi 0."""
    for c in cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df

def isi_template(wb_stream, nama, p1, p2, p3, pph21):
    """
    wb_stream: file-like (template xlsx)
    Mengembalikan BytesIO berisi workbook yang sudah terisi.
    """
    wb = load_workbook(wb_stream)
    ws = wb.active  # pakai sheet aktif

    # Isi sel sesuai ketentuan
    ws["D26"] = nama
    ws["C11"] = p1
    ws["C12"] = p2
    ws["C13"] = p3

    total_honor = (p1 or 0) + (p2 or 0) + (p3 or 0)
    ws["C14"] = total_honor

    ws["C16"] = pph21 or 0
    ws["C18"] = total_honor - (pph21 or 0)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# Upload database & template
db_file = st.file_uploader("Upload Database (Excel/CSV)", type=["xlsx", "csv"])
template_file = st.file_uploader("Upload Template SPD (Excel)", type=["xlsx"])

if db_file and template_file:
    # Baca database
    try:
        if db_file.name.lower().endswith(".csv"):
            df = pd.read_csv(db_file)
        else:
            df = pd.read_excel(db_file)
    except Exception as e:
        st.error(f"Gagal membaca database: {e}")
        st.stop()

    # Validasi kolom
    if not all(col in df.columns for col in REQUIRED_COLS):
        st.error(
            "Kolom database harus mengandung: "
            + ", ".join(REQUIRED_COLS)
        )
        st.stop()

    # Pastikan kolom honor & PPH numerik
    df = coerce_numeric(
        df,
        [
            "Honorarium Persiapan UKOMNAS",
            "Honorarium Pemantauan Briefing UKOMNAS",
            "Honorarium Pelaksanaan UKOMNAS",
            "PPH21",
        ],
    )

    st.subheader("Preview Database")
    st.dataframe(df.head())

    # ========== Generate per nama ==========
    nama_terpilih = st.selectbox("Pilih Nama", df["Nama"].astype(str).unique())

    if st.button("🔄 Generate Template (Satu Nama)"):
        row = df[df["Nama"].astype(str) == str(nama_terpilih)].iloc[0]

        # Ambil angka
        p1 = float(row["Honorarium Persiapan UKOMNAS"])
        p2 = float(row["Honorarium Pemantauan Briefing UKOMNAS"])
        p3 = float(row["Honorarium Pelaksanaan UKOMNAS"])
        pph21 = float(row["PPH21"])

        # Simpan salinan stream template (karena load_workbook mengonsumsi stream)
        tbytes = io.BytesIO(template_file.getvalue())

        filled = isi_template(
            wb_stream=tbytes,
            nama=row["Nama"],
            p1=p1,
            p2=p2,
            p3=p3,
            pph21=pph21,
        )

        st.success(f"Template untuk {row['Nama']} berhasil dibuat!")
        st.download_button(
            label="📥 Download File Excel",
            data=filled,
            file_name=f"Template_{row['Nama']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ========== Generate semua (ZIP) ==========
    if st.button("📦 Generate Semua Template (ZIP)"):
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for _, row in df.iterrows():
                p1 = float(row["Honorarium Persiapan UKOMNAS"])
                p2 = float(row["Honorarium Pemantauan Briefing UKOMNAS"])
                p3 = float(row["Honorarium Pelaksanaan UKOMNAS"])
                pph21 = float(row["PPH21"])

                # untuk setiap iterasi gunakan salinan template baru
                tbytes = io.BytesIO(template_file.getvalue())

                filled = isi_template(
                    wb_stream=tbytes,
                    nama=row["Nama"],
                    p1=p1,
                    p2=p2,
                    p3=p3,
                    pph21=pph21,
                )

                # Nama file aman (tanpa karakter aneh)
                safe_name = str(row["Nama"]).replace("/", "-").replace("\\", "-")
                zf.writestr(f"Template_{safe_name}.xlsx", filled.getvalue())

        zip_buffer.seek(0)
        st.success("Semua template berhasil dibuat!")
        st.download_button(
            label="📥 Download ZIP Semua Template",
            data=zip_buffer,
            file_name="Semua_Template_SPD.zip",
            mime="application/zip",
        )
